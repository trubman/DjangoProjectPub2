from collections import namedtuple
from datetime import datetime
from django.contrib import messages
import json
import copy

from openpyxl import load_workbook, Workbook
from sert.forms import (
    MeltForm,
    KernelForm,
    SertForm,
    AttachmentForm,
)


class ImportManager:
    def __init__(self):
        self.input_file = None
        self.importer = Importer()
        self.converter = Converter()
        self.loader = Loader()
        self.errors_list = [] # (level, text,)

    def get_errors(self):
        return self.errors_list

    def set_file(self, input_file):
        self.importer.set_input_file(input_file)
        if self.importer.get_errors():
            self.errors_list += self.importer.get_errors()

        if not self.importer.get_fatal_error():
            self.converter.set_input_data(self.importer.get_result())
            # if self.converter.get_errors():
            #     self.errors_list += self.converter.get_errors()

            self.loader.set_input_data(self.converter.get_result())
            if self.loader.get_errors():
                self.errors_list += self.loader.get_errors()


class Importer:
    def __init__(self):
        self.input_file = None
        self.result_dict = {'sert': [], 'attach': [], 'melt': [], }

        self.wb = Workbook()
        self.fatal_error = False
        self.errors_list = []  # (level, text,)
        self.ws_data_dict = {'sert': None, 'attach': None, 'melt': None,}
        self.ws_names_dict = {'sert': 'SERT', 'attach': 'ATTACH', 'melt': 'MELT',}
        self.ws_list = ['sert', 'attach', 'melt',]

        self.sert_names_conv = {
            'номер_СПГ': 'number_spg',
            'обозначение': 'designation',
            'наименование': 'denomination',
            'количество': 'quantity',
            'тип_сертификата': 'sert_type',
            'дата': 'date',
            'гальв_дата': 'galvan_date',
            'есть_драг': 'is_drag_met',
            'на_печать': 'is_print',
            'это_атом': 'is_atom',
            'атомный_договор': 'atom_contract',
        }
        self.sert_row_names = [
            'number_spg',
            'designation',
            'denomination',
            'quantity',
            'sert_type',
            'date',
            'galvan_date',
            'is_drag_met',
            'is_print',
            'is_atom',
            'atom_contract',
        ]
        self.sert_nt = namedtuple('sert_row', self.sert_row_names)
        self.sert_col_names = {}

        self.attach_names_conv = {
            'тип_сертификата': 'sert_type',
            'номер_СПГ': 'number_spg',
            'обозначение': 'designation',
            'наименование': 'denomination',
            'количество': 'quantity',

            'гальв_мат': 'galvan_material',
            'гальв_ед_изм': 'galvan_units',

            'n': 'n_index',
            'a': 'a_index',
            'b': 'b_index',
            'a1': 'a1_index',
            'b1': 'b1_index',
            'a2': 'a2_index',
            'b2': 'b2_index',

            'номер_плавки': 'melt_number',
            'номер_мат': 'material_id',
            'год_плавки': 'melt_year',
            'паспорт_плавки': 'melt_passport',

            'мех_по_гост': 'is_by_gost_material_number',
            'ед_изм': 'item_units',
        }
        self.attach_row_names = [
            'sert_type',
            'number_spg',
            'designation',
            'denomination',
            'quantity',

            'galvan_material',
            'galvan_units',
            'n_index',
            'a_index',
            'b_index',
            'a1_index',
            'b1_index',
            'a2_index',
            'b2_index',

            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',

            'is_by_gost_material_number',
            'item_units',
        ]
        self.attach_nt = namedtuple('attach_row', self.attach_row_names)
        self.attach_col_names = {}

        self.melt_names_conv = {
            'номер_плавки': 'melt_number',
            'номер_мат': 'material_id',
            'год_плавки': 'melt_year',
            'паспорт_плавки': 'melt_passport',

            'углерод_C': 'carboneum',
            'марганец_Mn': 'manganum',
            'кремний_Si': 'silicium',
            'сера_S': 'sulfur',
            'фосфор_P': 'phosphorus',
            'хром_Cr': 'chromium',
            'молибден_Mo': 'molybdaenum',
            'никель_Ni': 'niccolum',
            'ниобий_Nb': 'niobium',
            'титан_Ti': 'titanium',
            'медь_Cu': 'cuprum',
            'магний_Mg': 'magnesium',
            'железо_Fe': 'ferrum',

            'предел_прочности': 'tensile_strength',
            'предел_текучести': 'yield_strength',
            'относительное_удлинение': 'relative_extension',
            'относительное_сужение': 'relative_narrowing',
            'ударная_вязкость': 'impact_strength',
            'ударная_вязкость_kcu-60': 'impact_strength_60KCU',
            'ударная_вязкость_kcv-60': 'impact_strength_60KCV',
            'твердость': 'hardness',
            'мкк': 'mkk',
        }
        self.melt_row_names = [
            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',

            'carboneum',
            'manganum',
            'silicium',
            'sulfur',
            'phosphorus',
            'chromium',
            'molybdaenum',
            'niccolum',
            'niobium',
            'titanium',
            'cuprum',
            'magnesium',
            'ferrum',

            'tensile_strength',
            'yield_strength',
            'relative_extension',
            'relative_narrowing',
            'impact_strength',
            'impact_strength_60KCU',
            'impact_strength_60KCV',
            'hardness',
            'mkk',
        ]
        self.melt_nt = namedtuple('melt_row', self.melt_row_names)
        self.melt_col_names = {}

    def set_input_file(self, input_file):
        self.input_file = input_file

        self.get_xlsx_object()
        self.get_sheets_objects()
        self.fill_result_lists()

    def get_result(self):
        return self.result_dict

    def get_errors(self):
        return self.errors_list

    def get_fatal_error(self):
        return self.fatal_error

    def get_xlsx_object(self):
        if self.input_file.name.endswith('.xlsx'):
            self.wb = load_workbook(self.input_file)
            error_text = f'Файл "{self.input_file.name}" успешно воспринят.'
            error = (messages.SUCCESS, error_text)
            self.errors_list.append(error)
        else:
            self.fatal_error = True
            file_name_list = self.input_file.name.split('.')
            error_text = (f'Требуется ".xlsx"! Прискорбно, но загруженный файл имеет формат '
                          f'".{file_name_list[len(file_name_list) - 1]}": ¯\\_(ツ)_/¯ Я не знаю что с ним делать...')
            error = (messages.ERROR, error_text)
            self.errors_list.append(error)

    def get_sheets_objects(self):
        for name in self.ws_list:
            try:
                ws_name = self.ws_names_dict[name]
                self.ws_data_dict[name] = self.wb[ws_name]
                error_text = f'Лист "{name}" xlsx-файла успешно прочитан.'
                error = (messages.SUCCESS, error_text)
                self.errors_list.append(error)
            except KeyError:
                self.fatal_error = True
                error_text = (f'xlsx-файл не содержит лист "{name}". '
                              f'дальнейшая обработка файла не имеет смысла.')
                error = (messages.ERROR, error_text)
                self.errors_list.append(error)

    def empty_checker(self, nt):
        is_empty = False
        li = []
        for emp in nt:
            if isinstance(emp, type(None)):
                li.append(emp)
        if len(nt) == len(li):
            is_empty = True
        return is_empty

    def str_to_bolean_converter(self, value):
        # print('str_to_bolean_converter', value)
        if str(value).lower().strip() == 'да':
            res_value = True
        elif str(value).lower().strip() == 'нет':
            res_value = False
        else:
            res_value = None
        # print('str_to_bolean_converter', res_value)
        return res_value

    def convert_ws_to_list(self, ws_name):
        WS_NT = namedtuple('WS_NT', [
            'data_dict_name',
            'names_conv',
            'col_names',
            'row_names',
            'nt',
        ])
        WS_CHOISE = {
            'sert': WS_NT(
                data_dict_name='sert',
                names_conv='sert_names_conv',
                col_names='sert_col_names',
                row_names='sert_row_names',
                nt='sert_nt',
            ),
            'attach': WS_NT(
                data_dict_name='attach',
                names_conv='attach_names_conv',
                col_names='attach_col_names',
                row_names='attach_row_names',
                nt='attach_nt',
            ),
            'melt': WS_NT(
                data_dict_name='melt',
                names_conv='melt_names_conv',
                col_names='melt_col_names',
                row_names='melt_row_names',
                nt='melt_nt',
            ),
        }
        meta_ws = WS_CHOISE[ws_name]
        boolean_fields_names = ['is_print', 'is_drag_met', 'is_atom', 'is_by_gost_material_number',]

        ws = self.ws_data_dict[meta_ws.data_dict_name]
        # обработай первую строку
        head_row = ws.iter_rows(min_row=1, max_row=1)
        for row in head_row:
            for cell in row:
                name = self.__getattribute__(meta_ws.names_conv)[str(cell.value)]
                self.__getattribute__(meta_ws.col_names)[name] = cell.column_letter
        # обработай прочие строки
        for row_num in range(2, len(tuple(ws.rows)) + 1):
            row_content = []
            row = ws.iter_rows(min_row=row_num, max_row=row_num)
            for a in row:
                for aa in a:
                    row_content.append(aa)
            di = {}
            for row_name_num in range(len(self.__getattribute__(meta_ws.row_names))):
                cell_name = self.__getattribute__(meta_ws.row_names)[row_name_num]
                val = ws[f'{self.__getattribute__(meta_ws.col_names)[cell_name]}{row_num}'].value
                if cell_name not in boolean_fields_names:
                    di[cell_name] = val
                else:
                    di[cell_name] = self.str_to_bolean_converter(val)
            nt = self.__getattribute__(meta_ws.nt)(**di)
            is_empty = self.empty_checker(nt)
            if not is_empty:
                self.result_dict[meta_ws.data_dict_name].append(nt)

    def fill_result_lists(self):
        if not self.fatal_error:
            for ws_name in self.ws_list:
                self.convert_ws_to_list(ws_name)


class Converter:
    def __init__(self):
        self.input_dict = None
        self.input_lists_names = ['sert', 'attach', 'melt', ]
        self.result_lists_names = ['kernel', 'sert', 'attachment', 'melt',]
        self.data_pair = {'kernel': 'sert', 'sert': 'sert', 'attachment': 'attach', 'melt': 'melt',}
        self.processing_result = {'kernel': [], 'sert': [], 'attachment': [], 'melt': [],}
        # self.second_processing_result = {'kernel': [], 'sert': [], 'attachment': [], 'melt': [],}
        # self.third_processing_result = {'kernel': [], 'sert': [], 'attachment': [], 'melt': [],}

        self.kernel_pattern = {
            'number_spg': None,
            'designation': None,
            'denomination': None,
            'quantity': 1,
            'is_atom': False,
            'atom_contract': None,
        }
        self.kernel_fields_names = [
            'number_spg',
            'designation',
            'denomination',
            'quantity',
            'is_atom',
            'atom_contract',
        ]
        self.kernel_load_fields = self.kernel_fields_names
        self.kernel_nt = namedtuple('kernel_nt', self.kernel_fields_names)
        self.sert_pattern = {
            'id': None,
            'is_print': True,
            'number_spg': None,
            'sert_type': 'НАСОС',
            'number_unique': None,
            'date': datetime.now().date(),
            'galvan_date': datetime.now().date(),
            'conclusion_type': None,
            'is_drag_met': False,
            'guarantee_type': None,
            'sign_type': None,
        }
        self.sert_fields_names = [
            'id',
            'is_print',
            'number_spg',
            'sert_type',
            'number_unique',
            'date',
            'galvan_date',
            'conclusion_type',
            'is_drag_met',
            'guarantee_type',
            'sign_type',
        ]
        self.sert_load_fields = [
            'is_print',
            'number_spg',
            'sert_type',
            'date',
            'galvan_date',
            'is_drag_met',
        ]
        self.sert_nt = namedtuple('sert_nt', self.sert_fields_names)
        self.attachment_pattern = {
            'number_spg': None,
            'number_unique': None,
            'sert_type': None,

            'designation': None,
            'denomination': None,
            'quantity': None,
            'item_units': None,

            'n_index': None,
            'a_index': None,
            'b_index': None,
            'a1_index': None,
            'b1_index': None,
            'a2_index': None,
            'b2_index': None,
            'is_one': None,
            'is_two': None,
            'is_three': None,
            'is_four': None,

            'melt_number': None,
            'material_id': None,
            'melt_year': None,
            'melt_passport': None,
            'is_cast': False,

            'is_by_gost_material_number': True,
            'galvan_material': None,
            'galvan_units': None,
        }
        self.attachment_fields_names = [
            'number_spg',
            'number_unique',
            'sert_type',

            'designation',
            'denomination',
            'quantity',
            'item_units',

            'n_index',
            'a_index',
            'b_index',
            'a1_index',
            'b1_index',
            'a2_index',
            'b2_index',
            'is_one',
            'is_two',
            'is_three',
            'is_four',

            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',
            'is_cast',

            'is_by_gost_material_number',
            'galvan_material',
            'galvan_units',
        ]
        self.attachment_load_fields = [
            'number_spg',
            'sert_type',

            'designation',
            'denomination',
            'quantity',
            'item_units',

            'n_index',
            'a_index',
            'b_index',
            'a1_index',
            'b1_index',
            'a2_index',
            'b2_index',

            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',

            'is_by_gost_material_number',
            'galvan_material',
            'galvan_units',
        ]
        self.attachment_nt = namedtuple('attachment_nt', self.attachment_fields_names)
        self.melt_pattern = {
            'melt_id': None,
            'melt_number': None,
            'material_id': None,
            'melt_year': None,
            'melt_passport': None,
            'by_gost_number': None,

            'carboneum': None,
            'manganum': None,
            'silicium': None,
            'sulfur': None,
            'phosphorus': None,
            'chromium': None,
            'molybdaenum': None,
            'niccolum': None,
            'niobium': None,
            'titanium': None,
            'cuprum': None,
            'magnesium': None,
            'ferrum': None,

            'tensile_strength': None,
            'yield_strength': None,
            'relative_extension': None,
            'relative_narrowing': None,
            'impact_strength': None,
            'impact_strength_60KCU': None,
            'impact_strength_60KCV': None,
            'hardness': None,
            'mkk': None,
        }
        self.melt_fields_names = [
            'melt_id',
            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',
            'by_gost_number',

            'carboneum',
            'manganum',
            'silicium',
            'sulfur',
            'phosphorus',
            'chromium',
            'molybdaenum',
            'niccolum',
            'niobium',
            'titanium',
            'cuprum',
            'magnesium',
            'ferrum',

            'tensile_strength',
            'yield_strength',
            'relative_extension',
            'relative_narrowing',
            'impact_strength',
            'impact_strength_60KCU',
            'impact_strength_60KCV',
            'hardness',
            'mkk',
        ]
        self.melt_load_fields = [
            'melt_number',
            'material_id',
            'melt_year',
            'melt_passport',

            'carboneum',
            'manganum',
            'silicium',
            'sulfur',
            'phosphorus',
            'chromium',
            'molybdaenum',
            'niccolum',
            'niobium',
            'titanium',
            'cuprum',
            'magnesium',
            'ferrum',

            'tensile_strength',
            'yield_strength',
            'relative_extension',
            'relative_narrowing',
            'impact_strength',
            'impact_strength_60KCU',
            'impact_strength_60KCV',
            'hardness',
            'mkk',
        ]
        self.melt_nt = namedtuple('melt_nt', self.melt_fields_names)


    def set_input_data(self, input_dict):
        self.input_dict = input_dict

        self.convert_to_model()

    def get_result(self):
        return self.processing_result

    # def get_errors(self):
    #     return self.errors_list

    def convert_to_model(self):
        for name in self.result_lists_names:
            self.convert_to_model_row(name)

    def convert_to_model_row(self, name):
        NT = namedtuple('NT', [
            'pattern', 'load_fields', 'nt',
        ])
        PARAM = {
            'kernel': NT(
                pattern='kernel_pattern',
                load_fields='kernel_load_fields',
                nt='kernel_nt',
            ),
            'sert': NT(
                pattern='sert_pattern',
                load_fields='sert_load_fields',
                nt='sert_nt',
            ),
            'attachment': NT(
                pattern='attachment_pattern',
                load_fields='attachment_load_fields',
                nt='attachment_nt',
            ),
            'melt': NT(
                pattern='melt_pattern',
                load_fields='melt_load_fields',
                nt='melt_nt',
            ),
        }
        data_list_name = self.data_pair[name]
        data_list = self.input_dict[data_list_name]
        for data_row in data_list:
            di = copy.deepcopy(self.__getattribute__(PARAM[name].pattern))
            # print('старт', di)
            fields_names = self.__getattribute__(PARAM[name].load_fields)
            for field_name in fields_names:
                if ((di[field_name] != data_row.__getattribute__(field_name))
                    or (not data_row.__getattribute__(field_name))):
                    val = data_row.__getattribute__(field_name)
                    val = self.field_str_data_cleaner(val)
                    val = self.field_data_type_convert(field_name, val)
                    di[field_name] = val
            row = self.__getattribute__(PARAM[name].nt).__call__(**di)
            self.processing_result[name].append(row)

    def field_str_data_cleaner(self, val):
        if isinstance(val, type('str')):
            val = val.strip()
        return val

    def field_data_type_convert(self, field_name, val):
        bool_list = ['is_atom', 'is_print', 'is_drag_met', 'is_by_gost_material_number',]
        if isinstance(val, type('str')):
            if field_name in bool_list:
                if val.lower() in ['да', 'нет', ]:
                    if val.lower() == 'да':
                        val = True
                    elif val.lower() == 'нет':
                        val = False
        date_list = ['date', 'galvan_date',]
        if isinstance(val, type(datetime(2024, 1, 1, 0, 0))):
            if field_name in date_list:
                val = val.date()
        melt_id_list = ['melt_number', 'material_id', 'melt_year', 'melt_passport',]
        if field_name in melt_id_list:
            if not isinstance(val, type(None)):
                val = str(val)
        return val


class Loader:
    def __init__(self):
        self.input_data = None
        self.fatal_error = False
        self.sert_error = False
        self.errors_list = []  # (level, text,)
        self.kernel_data = []
        self.sert_data = []
        self.attachment_data = []
        self.melt_data = []
        self.saved_serts_list = []

    def set_input_data(self, input_data):
        self.input_data = input_data

        self.load_pre_data()
        self.do_load()

    def get_errors(self):
        return self.errors_list

    def load_pre_data(self):
        try:
            self.kernel_data = self.input_data['kernel']
            self.sert_data = self.input_data['sert']
            self.attachment_data = self.input_data['attachment']
            self.melt_data = self.input_data['melt']
        except KeyError:
            self.fatal_error = True
            error_text = (f'Loader: Входные данные не содержат необходимую '
                          f'информацию для загрузки в MODELS')
            error = (messages.ERROR, error_text)
            self.errors_list.append(error)

    def do_load(self):
        if not self.fatal_error:
            # print('fatal_error = ', self.fatal_error)
            self.load_melt()
            self.load_kernel()
            self.load_sert()
            if not self.sert_error:
                self.load_attach()

    def convert_form_errors(self, form):
        error_text_list = []
        error_json = form.errors.as_json()
        error_dict = json.loads(error_json)
        for k, v in error_dict.items():
            for i in v:
                t = i['message']
                text = f'{t} [{k}] '
                error_text_list.append(text)
        return error_text_list

    def load_melt(self):
        success_list = []
        model_name = 'Отливка(Melt):'
        for melt in self.melt_data:
            form = MeltForm(melt._asdict())
            if form.is_valid():
                form.save()
                ft = form.cleaned_data['melt_id']
                melt_name = f'{ft}'
                success_list.append(melt_name)
            else:
                error_text_list = self.convert_form_errors(form)
                for m in error_text_list:
                    melt_name = f'{melt.melt_number}-{melt.material_id}-{melt.melt_year}-{melt.melt_passport}'
                    error_text = f'{model_name} {m} || {melt_name}'
                    error = (messages.WARNING, error_text)
                    self.errors_list.append(error)
        if success_list:
            et = ', '.join(success_list)
            error_text = f'{model_name} {len(success_list)} записей || {et} сохранена(ы).'
            error = (messages.SUCCESS, error_text)
            self.errors_list.append(error)

    def load_kernel(self):
        success_list = []
        model_name = 'Продукт(Kernel):'
        for kernel in self.kernel_data:
            form = KernelForm(kernel._asdict())
            if form.is_valid():
                form.save()
                ft = form.cleaned_data['number_spg']
                kernel_name = f'{ft}'
                success_list.append(kernel_name)
            else:
                error_text_list = self.convert_form_errors(form)
                for m in error_text_list:
                    kernel_name = f'{kernel.number_spg}, {kernel.designation}, {kernel.denomination}'
                    error_text = f'{model_name} {m} || {kernel_name}'
                    error = (messages.WARNING, error_text)
                    self.errors_list.append(error)
        if success_list:
            et = ', '.join(success_list)
            error_text = f'{model_name} {len(success_list)} записей || {et} сохранен(ы).'
            error = (messages.SUCCESS, error_text)
            self.errors_list.append(error)

    def load_sert(self):
        success_list = []
        model_name = 'Сертификат(Sert):'
        for sert in self.sert_data:
            form = SertForm(sert._asdict())
            # print(sert._asdict())
            if form.is_valid():
                saved_sert = form.save()
                self.saved_serts_list.append(saved_sert)
                ft = form.cleaned_data['id']
                sert_name = f'{ft}'
                success_list.append(sert_name)
            else:
                self.sert_error = True
                error_text_list = self.convert_form_errors(form)
                for m in error_text_list:
                    sert_name = f'{sert.number_spg}-{sert.sert_type}'
                    error_text = f'{model_name} {m} || {sert_name}'
                    error = (messages.WARNING, error_text)
                    self.errors_list.append(error)
        if success_list:
            et = ', '.join(success_list)
            error_text = f'{model_name} {len(success_list)} записей || {et} сохранен(ы).'
            error = (messages.SUCCESS, error_text)
            self.errors_list.append(error)

    def load_attach(self):
        success_list = []
        model_name = 'Вложения(Attachment):'
        for attach in self.attachment_data:
            form = AttachmentForm(attach._asdict())
            if form.is_valid():
                form.save()
                if self.saved_serts_list:
                    for s in self.saved_serts_list:
                        ft1 = form.cleaned_data['number_spg'].number_spg
                        ft2 = form.cleaned_data['sert_type']
                        if s.id == f'{ft1}-{ft2}':
                            form.cleaned_data['number_unique'] = s.number_unique
                ft3 = form.cleaned_data['number_spg'].number_spg
                ft4 = form.cleaned_data['sert_type']
                attach_name = f'{ft3}-{ft4}'
                success_list.append(attach_name)
            else:
                error_text_list = self.convert_form_errors(form)
                for m in error_text_list:
                    attach_name = f'{attach.number_spg}-{attach.sert_type}'
                    error_text = f'{model_name} {m} || {attach_name}'
                    error = (messages.WARNING, error_text)
                    self.errors_list.append(error)
        if success_list:
            et = ', '.join(success_list)
            error_text = f'{model_name} {len(success_list)} записей || {et} сохранен(ы).'
            error = (messages.SUCCESS, error_text)
            self.errors_list.append(error)

